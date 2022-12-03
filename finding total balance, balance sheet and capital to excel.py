#Script for Upgrade Button: opening "שכר בכירים +ביצועי חברה" excel file.
#Fill the company preformance.

import openpyxl
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains
from decimal import Decimal
from selenium.webdriver.common.by import By 
from datetime import date

class UpdateFile:

    #function takes the string and returns the currence value. if it is Shekel it returns 1.
    def currenceTypeYearlyAndQuarterly(self, currence_type_title, excelQuarterlyDataPart, currence_yearly_dollar, currence_of_march_dollar ,currence_of_june_dollar, 
                     currence_of_september_dollar, currence_yearly_euro, currence_of_march_euro ,currence_of_june_euro , currence_of_september_euro):
        
        currence_part_title_1 = currence_type_title[currence_type_title.find("(")+1:currence_type_title.find(")")]
        currence_part_title_2 = currence_part_title_1.split(" ")
        currence_value = 0
        
        if  'דולרים' in currence_part_title_2:
            if currence_yearly_dollar != "future":
                
                currence_yearly_value_dollar = float(currence_yearly_dollar)
                if currence_of_september_dollar !="future":
                    if excelQuarterlyDataPart == 6:
                        currence_Quarterly_value_dollar = float(currence_of_september_dollar)
                else:
                     currence_Quarterly_value_dollar = currence_value
                     
                if currence_of_june_dollar !="future":                
                    if excelQuarterlyDataPart == 8:
                        currence_Quarterly_value_dollar = float(currence_of_june_dollar)
                else:
                    currence_Quarterly_value_dollar = currence_value
                    
                if currence_of_march_dollar !="future":                     
                    if excelQuarterlyDataPart == 10:
                        currence_Quarterly_value_dollar = float(currence_of_march_dollar)
                else:
                    currence_Quarterly_value_dollar = currence_value
                
                return currence_Quarterly_value_dollar, currence_yearly_value_dollar
            
            else:
                return currence_value
        
        elif 'ש"ח' in currence_part_title_2:
            
            currence_value = 1
            return currence_value  
        
        elif 'אירו' in currence_part_title_2:
            if currence_yearly_dollar != "future":
                currence_yearly_value_euro = float(currence_yearly_euro)
                if currence_of_september_euro !="future":
                    if excelQuarterlyDataPart == 6:
                        currence_Quarterly_value_euro = float(currence_of_september_euro)
                else:
                    currence_Quarterly_value_euro = currence_value     
                if currence_of_june_euro !="future":                
                    if excelQuarterlyDataPart == 8:
                        currence_Quarterly_value_euro = float(currence_of_june_euro)
                else:
                    currence_Quarterly_value_euro = currence_value
                    
                if currence_of_march_euro !="future":                     
                    if excelQuarterlyDataPart == 10:
                        currence_Quarterly_value_euro = float(currence_of_march_euro)     
                else:
                    currence_Quarterly_value_euro = currence_value
                
                return currence_Quarterly_value_euro, currence_yearly_value_euro
            else:
                return currence_value


    def currenceTypeYearly(self, currence_type_title, currence_yearly_dollar, currence_yearly_euro):
        
        currence_part_title_1 = currence_type_title[currence_type_title.find("(")+1:currence_type_title.find(")")]
        currence_part_title_2 = currence_part_title_1.split(" ")
        currence_value = 0
        
        if  'דולרים' in currence_part_title_2:
            if currence_yearly_dollar != "future":
                currence_yearly_value_dollar = float(currence_yearly_dollar)
                return currence_yearly_value_dollar
        
        elif 'ש"ח' in currence_part_title_2:
            
            currence_value = 1
            return currence_value  
        
        elif 'אירו' in currence_part_title_2:
            if currence_yearly_dollar != "future":
                currence_yearly_value_euro = float(currence_yearly_euro)
                return currence_yearly_value_euro


            
    
    #check if element exist if not return False alse return element.
    def driver_get_element_if_exists(self, driver,*args,**kwargs):
        try:
            element = driver.find_element(*args,**kwargs)
            return element
        except selenium.common.exceptions.NoSuchElementException as e:
            return False
    
    #takes take a string number with commas and returns an int with out the commas.
    def numberWithCommaToInt(self, total_balance):
        try:
            number = int(total_balance.replace(',', ''))
            return number
        except:
            return "no value"
    
    #the function takes the lable element and finds it's index in the list. 
    def elementIndex(self, element_lable, string_title):
        element_list=[]
        for i in element_lable :
            if (i.get_attribute("innerText") != "null") :
                element_list.append(i.get_attribute("innerText"))  
                
        for j in range(0, len(element_list),1):
            for substring_title in string_title:
                substring_length = len(substring_title)
                if 0 < len(element_list[j]) < 3*substring_length or substring_length == len(element_list[j]):
                    if substring_title in element_list[j]: 
                        return j
        return "element does not exist"
    
    
    
    #the function takes elements. generate the labels and from the labels returns excel column and title index.
    def titleIndexAndExcel(self, element_lable):
    
        element_list=[]
        for i in element_lable :
            if (i.get_attribute("innerText") != "null") :
                element_list.append(i.get_attribute("innerText"))  
                
        data_list = []    
        for j in range(0,len(element_list),1):
            title = element_list[j].split(" ")
            
            calendar = {"שנתי":3,"מרץ":10,"יוני":8,"ספטמבר":6} 
            
            for month in title:
                for calendar_key, calendar_value in calendar.items():
                    if month == calendar_key:
                        excelColumnMonthValue = calendar_value
            
            todays_date = date.today()
            current_year = todays_date.year 
            one_year_before = current_year - 1
            two_years_before = current_year - 2
            
            what_month_is_it = todays_date.month
            
            if 3 <= what_month_is_it <= 12:
                new_year = 0
            else:
                new_year = 1
                
            for title_part in title:
                if str(current_year) in title_part:
                    year_diff = 0
                elif str(one_year_before) in title_part:
                    year_diff = -1
                elif str(two_years_before) in title_part:
                    year_diff = -2
                else:
                    year_diff = False
            
            if (year_diff == 0 and new_year == 0) and (excelColumnMonthValue == 6 or excelColumnMonthValue == 8 or excelColumnMonthValue == 10):
                data_list.append((excelColumnMonthValue,j+1))
            elif (year_diff == -1 and new_year == 0) and excelColumnMonthValue == 3:
                data_list.append((excelColumnMonthValue,j+1))
            elif (year_diff == -1 and new_year == 1) and (excelColumnMonthValue == 6 or excelColumnMonthValue == 8 or excelColumnMonthValue == 10):
                data_list.append((excelColumnMonthValue,j+1))
            elif year_diff == -2 and new_year == 1 and excelColumnMonthValue == 3:
                data_list.append((excelColumnMonthValue,j+1))
            else:
                if year_diff == False:
                    return 'data from old years'
            
        if len(data_list) > 0:
            if len(data_list) > 2:
                del data_list[1]
                return data_list
            else:
                return data_list
        else:    
            return "title does not exist"    
    
    
    #main function, automation and stuff   
    def main(self, PATH, ExcelSourcePath, currence_yearly_dollar, currence_of_march_dollar ,currence_of_june_dollar, 
                     currence_of_september_dollar, currence_yearly_euro, currence_of_march_euro ,currence_of_june_euro , currence_of_september_euro): 

         driver = webdriver.Chrome(PATH)
         book = openpyxl.load_workbook(ExcelSourcePath)
         sheet = book["ביצועי חברות"]
         
         number_of_rows = len(sheet['A'])
         for securities_number in range(2,number_of_rows,1):
             company_securities_number = sheet.cell(row= securities_number, column=2).value
             if company_securities_number == "n.a.":
                 pass
             else:
                 #go to tase site.
                 driver.get("http://www.tase.co.il")
                 driver.maximize_window()
                 time.sleep(2)
                 #find the search and inset the company_securities_number.
                 element = driver.find_element_by_id("headerSearch")
                 element.send_keys(company_securities_number)
                 time.sleep(2)   
                 
                 #checking if the are no result in the list
                 is_element = self.driver_get_element_if_exists(driver,By.XPATH,"//div[@class='col-sm-9']")
                 if is_element:
                     element = driver.find_element_by_xpath("//div[@class='col-sm-9']")
                     no_result = element.text
                     if no_result == 'לא נמצאו תוצאות העונות לחיפוש שלך':
                         continue
                 
                 #find the name of the company in the list and click on it(comparing strng_title and webElements).
                 elements = driver.find_elements_by_xpath("//a[@class='row no_margin']/div[@class='col-sm-8']/div")
                 elements_lable = driver.find_elements_by_xpath("//span[@class='search_result_type']")
                 string_title = ['מניות','יחידת השתתפות']
                 is_element = self.elementIndex(elements_lable, string_title)
                 if is_element == "element does not exist":
                     continue
                 
                 elements[self.elementIndex(elements_lable, string_title)].click()
                 time.sleep(5)
                 
                 #if there is no data
                 is_element = self.driver_get_element_if_exists(driver,By.XPATH,"//div[@class='col-xs-12']")
                 if is_element:
                     is_element = driver.find_element_by_xpath("//div[@class='col-xs-12']")
                     error = is_element.text
                     error_string = error.split(" ")
                     if "null" in error_string:
                         continue
                 
                 #find the short annual report link. scroll down to it and click on it.
                 element = driver.find_element_by_xpath("//span[contains(text(),'לתמצית דוחות כספיים')]")
                 #actions = ActionChains(driver)
                 #actions.move_to_element(element).perform()
                 desired_y = (element.size['height'] / 2) + element.location['y']
                 current_y = (driver.execute_script('return window.innerHeight') * 4/5) + driver.execute_script('return window.pageYOffset')
                 scroll_y_by = desired_y - current_y
                 driver.execute_script("window.scrollBy(0, arguments[0]);", scroll_y_by)
                 element.click()
                 time.sleep(5)
                 #if there are no results on page
                 is_element = self.driver_get_element_if_exists(driver,By.XPATH,"//div[@class='table_page_noresults']")
                 if is_element:
                     element = driver.find_element_by_xpath("//div[@class='table_page_noresults']")
                     if element.text == 'לא נמצאו דיווחים בנושא':
                         continue
                 
                 
                 #checking how many columns there are in the table
                 elements = driver.find_elements_by_xpath("//th[@class='ColW_10']")
                 how_many_columns = len(elements)
                 if how_many_columns == 3:
                     #returns (x,y). where x is the excel value and y is the column value 
                     Data = self.titleIndexAndExcel(elements)  
                     if Data =='data from old years' or Data == "title does not exist":
                         continue
                     else:
                         quarterlyData, yearlyData = Data
                                           
                         excelQuarterlyDataPart,titleIndexQuarterlyData = quarterlyData
                         excelYearlyDataPart,titleIndexYearlyDataPart = yearlyData 
                         
                         #annal report data title: dollars ,shekel or uro
                         element = driver.find_element_by_xpath("//div[@class='general_popover_inner']")
                         currence_type_title = element.text
                         currence = self.currenceTypeYearlyAndQuarterly(currence_type_title, excelQuarterlyDataPart, currence_yearly_dollar, currence_of_march_dollar ,currence_of_june_dollar, 
                                          currence_of_september_dollar, currence_yearly_euro, currence_of_march_euro ,currence_of_june_euro , currence_of_september_euro)
                         
                         if currence == 1 or currence == 0:   
                             currence_quarterly_value = currence
                             currence_yearly_value = currence      
                         else:
                             currence_quarterly_value, currence_yearly_value = currence
                         #elements_label are the subjects elements
                         elements_label = driver.find_elements_by_xpath("//td[@class='ColW_9']")
                         #elemenets are the elements table
                         elements = driver.find_elements_by_xpath("//td[@class='ColW_10 dir_ltr']")
                         
                         string_title = ['סך מאזן','סה"כ נכסים']
                         rowIndexInSite = self.elementIndex(elements_label, string_title)
                         total_balance_quarterly = elements[3*titleIndexQuarterlyData*rowIndexInSite].text
                         total_balance_yearly = elements[3*rowIndexInSite+2].text
                         
                         string_title = ['הון למאזן']
                         rowIndexInSite = self.elementIndex(elements_label, string_title)
                         capital_to_balance_sheet_quarterly = elements[3*titleIndexQuarterlyData*rowIndexInSite].text
                         capital_to_balance_sheet_yearly = elements[3*rowIndexInSite+2].text
                         
                         string_title = ['תשואה להון']
                         rowIndexInSite = self.elementIndex(elements_label, string_title)
                         return_on_equity = elements[3*rowIndexInSite+2].text
    
                         total_balance_without_comma_quarterly = self.numberWithCommaToInt(total_balance_quarterly)
                         total_balance_without_comma_yearly = self.numberWithCommaToInt(total_balance_yearly)
                         
                             
                         if total_balance_without_comma_quarterly == "no value":
                             
                             if excelQuarterlyDataPart == 6:
                                back_Ground_Color="FF7F50"
                             elif excelQuarterlyDataPart == 8:    
                                back_Ground_Color="6495ED"
                             elif excelQuarterlyDataPart == 10:    
                                back_Ground_Color="DFFF00"
                                 
                             my_color = openpyxl.styles.colors.Color(rgb=back_Ground_Color)
                             my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                             
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart).fill = my_fill
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart).value = total_balance_without_comma_quarterly
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart+1).fill = my_fill
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart+1).value = capital_to_balance_sheet_quarterly
                         else:
                             if excelQuarterlyDataPart == 6:
                                back_Ground_Color="FF7F50"
                             elif excelQuarterlyDataPart == 8:    
                                back_Ground_Color="6495ED"
                             elif excelQuarterlyDataPart == 10:    
                                back_Ground_Color="DFFF00"
                                
                             my_color = openpyxl.styles.colors.Color(rgb=back_Ground_Color)
                             my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                             
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart).fill = my_fill
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart).value = total_balance_without_comma_quarterly*currence_quarterly_value
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart+1).fill = my_fill
                             sheet.cell(row= securities_number, column=excelQuarterlyDataPart+1).value = capital_to_balance_sheet_quarterly
                         
                         if total_balance_without_comma_yearly == "no value":
                             my_color = openpyxl.styles.colors.Color(rgb="DE3163")
                             my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                             
                             sheet.cell(row= securities_number, column=excelYearlyDataPart).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart).value = total_balance_without_comma_yearly
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+1).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+1).value = capital_to_balance_sheet_yearly
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+2).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+2).value = return_on_equity
                         else:
                             my_color = openpyxl.styles.colors.Color(rgb="DE3163")
                             my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                             
                             sheet.cell(row= securities_number, column=excelYearlyDataPart).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart).value = total_balance_without_comma_yearly*currence_yearly_value
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+1).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+1).value = capital_to_balance_sheet_yearly
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+2).fill = my_fill
                             sheet.cell(row= securities_number, column=excelYearlyDataPart+2).value = return_on_equity


                 elif how_many_columns == 2:
                     element = driver.find_elements_by_xpath("//td[@class='ColW_10 dir_ltr']")
                     
                     total_balance = element[2].text
                     capital_to_balance_sheet = element[44].text
                     return_on_equity = element[46].text
                     
                     #returns (x,y). where x is the excel value and y is the column value 
                     Data = self.titleIndexAndExcel(elements)  
                     if Data =='data from old years' or Data == "title does not exist":
                         continue
                     else:
                         for yearlyData in Data:
                             excelYearlyDataPart,titleIndexYearlyDataPart = yearlyData 
                             
                             #annal report data title: dollars ,shekel or uro
                             element = driver.find_element_by_xpath("//div[@class='general_popover_inner']")
                             currence_type_title = element.text
                             currence = self.currenceTypeYearly(currence_type_title, currence_yearly_dollar,currence_yearly_euro)
                             
                             if currence == 1 or currence == 0:   
                                 currence_yearly_value = currence      
                             else:
                                 currence_yearly_value = currence
                             
                             total_balance_without_comma_yearly = self.numberWithCommaToInt(total_balance)
                             if total_balance_without_comma_yearly == "no value":
                                my_color = openpyxl.styles.colors.Color(rgb="DE3163")
                                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                                
                                sheet.cell(row= securities_number, column=excelYearlyDataPart).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart).value = total_balance_without_comma_yearly
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+1).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+1).value = capital_to_balance_sheet
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+2).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+2).value = return_on_equity
                             else:
                                my_color = openpyxl.styles.colors.Color(rgb="DE3163")
                                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
                                
                                sheet.cell(row= securities_number, column=excelYearlyDataPart).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart).value = total_balance_without_comma_yearly*currence_yearly_value
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+1).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+1).value = capital_to_balance_sheet
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+2).fill = my_fill
                                sheet.cell(row= securities_number, column=excelYearlyDataPart+2).value = return_on_equity

             book.save(ExcelSourcePath)
